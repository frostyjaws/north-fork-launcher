
import streamlit as st
import pandas as pd
import openpyxl
import json
import os
import requests
import re

st.set_page_config(page_title="North Fork Launcher", layout="centered")

CREDENTIALS_FILE = "stored_credentials.json"
st.title("North Fork Launcher")

def load_credentials():
    if os.path.exists(CREDENTIALS_FILE):
        with open(CREDENTIALS_FILE, "r") as f:
            return json.load(f)
    return {}

def save_credentials(data):
    with open(CREDENTIALS_FILE, "w") as f:
        json.dump(data, f)

creds = load_credentials()

with st.expander("Enter & Save API Credentials"):
    st.subheader("Shopify")
    shop_domain = st.text_input("Shopify Store URL", value=creds.get("shopify_domain", ""))
    shop_token = st.text_input("Shopify Admin API Token", type="password", value=creds.get("shopify_token", ""))

    st.subheader("Amazon")
    seller_id = st.text_input("Amazon Seller ID", value=creds.get("seller_id", ""))
    marketplace_id = st.text_input("Marketplace ID", value=creds.get("marketplace_id", "ATVPDKIKX0DER"))
    client_id = st.text_input("Amazon LWA Client ID", value=creds.get("client_id", ""))
    client_secret = st.text_input("Amazon LWA Client Secret", type="password", value=creds.get("client_secret", ""))
    refresh_token = st.text_input("Refresh Token", type="password", value=creds.get("refresh_token", ""))

    if st.button("Save Credentials"):
        creds = {
            "shopify_domain": shop_domain,
            "shopify_token": shop_token,
            "seller_id": seller_id,
            "marketplace_id": marketplace_id,
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token
        }
        save_credentials(creds)
        st.success("Credentials saved successfully!")

with st.expander("Test API Credentials"):
    if st.button("Test Shopify"):
        if shop_domain and shop_token:
            url = f"https://{shop_domain}/admin/api/2023-07/shop.json"
            headers = {"X-Shopify-Access-Token": shop_token}
            try:
                res = requests.get(url, headers=headers)
                if res.status_code == 200:
                    st.success("Shopify credentials are valid!")
                else:
                    st.error(f"Shopify error: {res.status_code} - {res.text}")
            except Exception as e:
                st.error(f"Error connecting to Shopify: {e}")
        else:
            st.warning("Missing Shopify credentials.")

    if st.button("Test Amazon"):
        if client_id and client_secret and refresh_token:
            try:
                token_url = "https://api.amazon.com/auth/o2/token"
                data = {
                    "grant_type": "refresh_token",
                    "refresh_token": refresh_token,
                    "client_id": client_id,
                    "client_secret": client_secret
                }
                res = requests.post(token_url, data=data)
                if res.status_code == 200:
                    st.success("Amazon SP-API credentials are valid!")
                else:
                    st.error(f"Amazon error: {res.status_code} - {res.text}")
            except Exception as e:
                st.error(f"Error connecting to Amazon: {e}")
        else:
            st.warning("Missing Amazon credentials.")

dry_run = st.checkbox("DRY RUN (no live upload)", value=True)
uploaded_file = st.file_uploader("Upload your mockup PNG", type=["png"])

if uploaded_file:
    filename_raw = uploaded_file.name.replace("-mockup.png", "").replace(".png", "").strip()
    filename_base = filename_raw.replace(" ", "").lower()
    display_title = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', filename_raw).title().strip()

    image_url = f"https://cdn.shopify.com/s/files/1/placeholder/{filename_base}.png"
    st.success(f"Mock image URL: {image_url}")

    # Shopify CSV update
    try:
        shopify_df = pd.read_csv("templates/Baby Onesie Shopify Flat File.csv")
        for i in range(1, 12):
            shopify_df.loc[i, 'Handle'] = f"{filename_base}baby-bodysuit-clothes-bodysuit-newborn"
            shopify_df.loc[i, 'Title'] = f"{display_title} - Baby Bodysuit Clothes Bodysuit Newborn"
        shopify_df.loc[1, 'Image Src'] = image_url
        shopify_df.loc[2, 'Image Src'] = "https://cdn.shopify.com/s/files/1/0545/2018/5017/files/12efccc074d5a78e78e3e0be1150e85c5302d855_aa77eee7-fccd-4fab-85f0-338ada8776b7.jpg?v=1743071089"
        output_name = f"{filename_base}_Shopify.csv"
        shopify_df.to_csv(output_name, index=False)
        with open(output_name, "rb") as f:
            st.download_button("Download Shopify CSV", f, file_name=output_name)
    except Exception as e:
        st.error(f"Shopify CSV error: {e}")

    # Amazon flat file update
    try:
        wb = openpyxl.load_workbook("templates/Baby Onesie Amazon Flat File.xlsx")
        ws = wb.active
        ws["A4"] = "leotard"
        ws["B4"] = f"{filename_base}-Parent"
        ws["E4"] = f"{display_title} - Baby Bodysuit Clothes Bodysuit Newborn"
        for row in range(5, 32):
            suffix = f"Var{row-4}"
            ws[f"A{row}"] = "leotard"
            ws[f"B{row}"] = f"{filename_base}-{suffix}"
            ws[f"E{row}"] = f"{display_title} - Baby Bodysuit Clothes Bodysuit Newborn"
            ws[f"T{row}"] = image_url
        amazon_name = f"{filename_base}_Amazon.xlsx"
        wb.save(amazon_name)
        with open(amazon_name, "rb") as f:
            st.download_button("Download Amazon Flat File", f, file_name=amazon_name)
    except Exception as e:
        st.error(f"Amazon flat file error: {e}")

    if st.button("Submit to Shopify + Amazon"):
        if dry_run:
            st.warning("DRY RUN active — no live submission sent.")
        else:
            st.success("Live submission would be performed here.")
else:
    st.info("Upload a PNG to begin.")
